# from pathlib import Path
# import openpyxl
# import datetime

# path_file = Path(__file__).parent / '装修费用记录.xlsx'
# # 加载工作薄
# wb = openpyxl.load_workbook(path_file)
# #  wb.sheetnames 获取工作薄中的sheet页名字
# sheetnames = wb.sheetnames
# print(sheetnames)

# # 获取工作表，传入的sheet页为列表
# sheet = wb.worksheets[0]
# # 获取当前工作表的范围
# print(sheet.dimensions)
# # 获取当前工作表的行数和列数
# print(sheet.max_row, sheet.max_column)

# sheet1 = wb.worksheets[1]   
# print(sheet1.dimensions)
# print(sheet1.max_row, sheet1.max_column)


# # 获取指定单元格的值
# print(sheet.cell(3, 5).value)
# print(sheet['F2'].value)
# print(sheet['G77'].value)

# # 获取多个单元格的值（嵌套元组）
# print(sheet['A2:F5'])


# # 读取所有单元格的数据
# for row_ch in range(2, sheet.max_row + 1):
#     for col_ch in 'ABCDEFGH':
#         value = sheet[f'{col_ch}{row_ch}'].value
#         print(value)
#         if type(value) == datetime.datetime:
#             print(value.strftime(('%Y年%m月%d日'), end='\t'))
#         elif type(value) == int:
#             print(f'{value:<10d}', end='\t')
#         elif type(value) == float:
#             print(f'{value:.4f}', end='\t')
#         else:
#             print(value, end='\t')
#     print() 

        
# 写入Excel 操作
# import random
# import openpyxl
# from pathlib import Path


# # 第一步 创建工作薄
# wb = openpyxl.Workbook()

# # 第二步 添加工作表
# sheet = wb.active
# sheet.title = '期末成绩'

# titles = ('姓名', '语文', '数学', '英语')
# for col_index, title in enumerate(titles):
#     sheet.cell(1, col_index + 1, title)

# names = ('关羽', '张飞', '赵云', '马超', '黄忠')
# for row_index, name in enumerate(names):
#     sheet.cell(row_index + 2, 1, name)
#     for col_index in range(2, 5):
#         sheet.cell(row_index + 2, col_index, random.randrange(50, 101))


# save_path = Path(__file__).parent / '考试成绩表.xlsx'
# wb.save(save_path)

# # wb.save('考试成绩表.xlsx')


# from pathlib import Path
# import openpyxl

# file_path = Path(__file__).parent / '装修费用记录.xlsx'
# wb = openpyxl.load_workbook(file_path)
# print(wb.sheetnames)
# # 获取工作表名称
# print(wb.sheetnames[1])
# # 获取工作表对象
# print(wb.worksheets[1])
# # 先拿到工作表对象，再访问它的属性
# sheet = wb.worksheets[1]
# print(sheet.dimensions)
# # 获取行列数，多少行多少列
# print(sheet.max_column)
# print(sheet.max_row)
# # 读取指定单元格的值，行列索引从1开始
# print(sheet.cell(1,1).value)

# 读取 装修费用记录.xlsx，输出这个工作簿中所有工作表的名称。
# from pathlib import Path
# from openpyxl import load_workbook

# file_path = Path(__file__).parent / '装修费用记录.xlsx'
# wb = load_workbook(file_path)
# print(wb.sheetnames)
# # 获取第 1 个工作表，输出它的 dimensions、max_row 和 max_column。
# sheet = wb.worksheets[0]
# print(sheet.dimensions)
# print(f'{sheet.max_column} {sheet.max_row}')
# # 分别用两种方式读取单元格 B2 的值：
# # sheet.cell(2, 2).value 和 sheet['B2'].value
# print(sheet.cell(2,2).value)
# print(sheet['B2'].value)
# # 读取区域 A1:C3 的所有单元格，并用双重循环打印它们的值。
# print(sheet['A1:C3'])
# for row in sheet['A1:C3']:
#     for cell in row:
#         print(cell.value)
# # 遍历某个工作表的前 5 行、前 4 列，把每个单元格内容输出出来。
# # cell(row, column) 先行后列
# for row in range(1,6):
#     for col in range(1,5):
#         print(sheet.cell(row,col).value)



# 提高题
# 6. 新建一个 Excel 文件 学生成绩.xlsx，创建工作表 成绩表，写入表头：姓名、语文、数学、英语，再写入 5 个学生的成绩。

from openpyxl import Workbook
from openpyxl import load_workbook
from pathlib import Path
import random


wb = Workbook()
sheet = wb.active
sheet.title = '成绩表'

titles = ('姓名', '语文', '数学', '英语')
for col_index, title in enumerate(titles):
    sheet.cell(1, col_index + 1, title)


names = ('关羽', '张飞', '赵云', '马超', '黄忠')
for row_index, name in enumerate(names):
    sheet.cell(row_index + 2, 1, name)
    for col_index in range(2, 5):
        sheet.cell(row_index + 2, col_index, random.randrange(50, 101))



# 7. 在 学生成绩.xlsx 中新增一列 总分，用公式计算每个学生三门课的总成绩。
sheet['E1'] = '总分'
for row_index in range(2,7):
    sheet[f'E{row_index}'] = f'=SUM(B{row_index}:D{row_index})'
# 8. 再新增一列 平均分，用公式计算每个学生的平均成绩。
sheet['F1'] = '平均分'
for row_index in range(2,7):
    sheet[f'F{row_index}'] = f'=AVERAGE(B{row_index}:D{row_index})'
# 9. 把表头设置成加粗、居中、字体颜色为蓝色，并把第 1 行行高调大一点。
# 10. 把 平均分 这一列的列宽调整得更合适一些。

# 综合题
# 11. 统计 成绩表 中语文、数学、英语三门课的平均成绩，并把结果写到表格最后一行。
# 12. 根据三门课的成绩数据，生成一个柱状图，插入到工作表中合适的位置。
# 13. 读取一个已有 Excel 文件，把其中“金额”这一列的数据全部取出，计算总和后输出。
sheet['G1'] = '金额'
for row_index in range(2, 7):
    sheet.cell(row_index, 7, random.randrange(10,50))

wb.save('学生成绩.xlsx')

file_path = Path(__file__).parent.parent.parent / '学生成绩.xlsx'
wb1 = load_workbook(file_path)
print(wb1.sheetnames)
sheet1 = wb1.worksheets[0]
print(sheet1.max_row)
print(sheet1.max_column)
for row_index in range(1, sheet1.max_row + 1):
    print(sheet1.cell(row_index, 7).value)
# 14. 遍历整个工作表，找出所有分数低于 60 分的学生姓名和科目。

for row_index in range(2, 7):
    student_name = sheet1.cell(row_index, 1).value
    for col_index in range(2, 5):
        score = sheet1.cell(row_index,col_index).value 
        subject = sheet1.cell(1, col_index).value
        if score < 60:
            print(f"{student_name}的 {subject}低于60分。{score}")
# 15. 自己设计一个 家庭开支表.xlsx，至少包含日期、项目、金额三列，并完成：
# - 写入数据
# - 计算总支出
# - 设置表头样式
# - 插入一个简单图表

